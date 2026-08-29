"""
Data Processor Module
Handles all Excel processing steps:
1. Bot output → Suppression sheet
2. Filter in-stock / correctly reflecting ASINs
3. Vlookup sellable qty from inventory
4. Remove < 30 qty
5. Compare with master suppression file → find new suppressions
6. Update master file
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# STEP 1: Build suppression dataframe from bot output
# ─────────────────────────────────────────────
def build_suppression_sheet(bot_results):
    """
    Takes bot scrape results list → builds suppression dataframe
    Adds check(ASIN=ASIN Reflect) column
    """
    df = pd.DataFrame(bot_results, columns=['ASIN', 'Availability', 'Asin Search', 'Asin Refelct'])
    df['check(ASIN=ASIN Reflect)'] = df['ASIN'] == df['Asin Refelct']
    logger.info(f"Step 1 complete: {len(df)} total ASINs in bot output")
    return df


# ─────────────────────────────────────────────
# STEP 2: Remove in-stock / correctly reflecting ASINs
# ─────────────────────────────────────────────
def filter_suppressed_asins(df):
    """
    Remove rows where:
    - Availability != 'Currently unavailable.' AND Asin Search == 'Yes' AND check == True
    (These are fine / not suppressed — we don't need them)
    Keeps only potentially suppressed ASINs
    """
    before = len(df)

    # Rows to REMOVE = in-stock + asin search yes + check true
    mask_to_remove = (
        (~df['Availability'].str.contains('Currently unavailable', case=False, na=False)) &
        (df['Asin Search'].str.strip().str.lower() == 'yes') &
        (df['check(ASIN=ASIN Reflect)'] == True)
    )

    df_filtered = df[~mask_to_remove].copy().reset_index(drop=True)
    removed = before - len(df_filtered)
    logger.info(f"Step 2 complete: Removed {removed} in-stock/healthy ASINs. Remaining: {len(df_filtered)}")
    return df_filtered


# ─────────────────────────────────────────────
# STEP 3: Vlookup sellable qty from inventory report
# ─────────────────────────────────────────────
def get_latest_month_sheet(inventory_path):
    """Get the latest month's sheet from the inventory report"""
    xl = pd.ExcelFile(inventory_path)
    sheets = xl.sheet_names

    # Filter sheets that match 'Master File_' pattern
    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    master_sheets = [s for s in sheets if 'Master File' in s]

    if not master_sheets:
        raise ValueError("No 'Master File_' sheets found in inventory report!")

    # Find the latest month
    def month_key(sheet_name):
        for i, m in enumerate(month_order):
            if m in sheet_name:
                return i
        return -1

    latest_sheet = max(master_sheets, key=month_key)
    logger.info(f"Using inventory sheet: {latest_sheet}")
    return latest_sheet


def vlookup_sellable_qty(df_suppressed, inventory_path):
    """
    VLookup Sellable Qty from inventory report
    Removes N/A rows (ASINs not in inventory)
    """
    latest_sheet = get_latest_month_sheet(inventory_path)
    df_inv = pd.read_excel(inventory_path, sheet_name=latest_sheet)

    # Make sure ASIN column exists
    if 'ASIN' not in df_inv.columns:
        raise ValueError(f"'ASIN' column not found in inventory sheet '{latest_sheet}'")

    if 'Sellable Qty' not in df_inv.columns:
        raise ValueError(f"'Sellable Qty' column not found in inventory sheet '{latest_sheet}'")

    # Build lookup dict: ASIN → Sellable Qty
    sellable_map = dict(zip(df_inv['ASIN'].astype(str).str.strip(),
                            df_inv['Sellable Qty']))

    # Map sellable qty
    df_suppressed = df_suppressed.copy()
    df_suppressed['sellable'] = df_suppressed['ASIN'].astype(str).str.strip().map(sellable_map)

    before = len(df_suppressed)
    # Remove rows where sellable is N/A (not in inventory)
    df_suppressed = df_suppressed.dropna(subset=['sellable']).copy()
    df_suppressed['sellable'] = df_suppressed['sellable'].astype(int)
    removed = before - len(df_suppressed)

    logger.info(f"Step 3 complete: Removed {removed} ASINs not in inventory. Remaining: {len(df_suppressed)}")
    return df_suppressed


# ─────────────────────────────────────────────
# STEP 4: Remove ASINs with sellable qty < 30
# ─────────────────────────────────────────────
def remove_low_inventory(df, min_qty=30):
    """Remove ASINs with sellable qty less than min_qty"""
    before = len(df)
    df_filtered = df[df['sellable'] >= min_qty].copy().reset_index(drop=True)
    removed = before - len(df_filtered)
    logger.info(f"Step 4 complete: Removed {removed} ASINs with sellable < {min_qty}. Remaining: {len(df_filtered)}")
    return df_filtered


# ─────────────────────────────────────────────
# STEP 5: Find NEW suppressions vs master file
# ─────────────────────────────────────────────
def find_new_suppressions(df_current, master_path, master_sheet):
    """
    Compare current ASINs against master Suppression_Sheet__1_
    Returns DataFrame of NEW suppressed ASINs (not in master file)
    """
    df_master = pd.read_excel(master_path, sheet_name=master_sheet)

    if 'ASIN' not in df_master.columns:
        raise ValueError(f"'ASIN' column not found in master sheet '{master_sheet}'")

    existing_asins = set(df_master['ASIN'].astype(str).str.strip().tolist())
    current_asins = df_current['ASIN'].astype(str).str.strip().tolist()

    new_asins = [a for a in current_asins if a not in existing_asins]
    df_new = df_current[df_current['ASIN'].isin(new_asins)].copy().reset_index(drop=True)

    logger.info(f"Step 5 complete: Found {len(df_new)} NEW suppressed ASINs (not in master)")
    return df_new, df_master


# ─────────────────────────────────────────────
# STEP 6: Update master suppression file
# ─────────────────────────────────────────────
def update_master_file(df_new, master_path, master_sheet):
    """
    Add new suppressed ASINs to master Suppression_Sheet__1_
    with today's date as Suppressed Date
    """
    if len(df_new) == 0:
        logger.info("No new ASINs to add to master file.")
        return

    df_master = pd.read_excel(master_path, sheet_name=master_sheet)

    today = datetime.now().strftime("%d-%m-%Y")
    new_rows = []
    for _, row in df_new.iterrows():
        new_rows.append({
            'Suppressed Date': today,
            'Live Date': '',
            'ASIN': row['ASIN'],
            'GL': '',
            'Sellable Inventory': row.get('sellable', ''),
            'Unsellable inventory': '',
            'Remark': '',
            'Status': 'Suppressed',
            'New ASIN': ''
        })

    df_new_rows = pd.DataFrame(new_rows)
    df_updated = pd.concat([df_master, df_new_rows], ignore_index=True)

    # Write back to the specific sheet using openpyxl to preserve other sheets
    wb = openpyxl.load_workbook(master_path)

    # Remove existing sheet and recreate
    if master_sheet in wb.sheetnames:
        del wb[master_sheet]

    ws = wb.create_sheet(master_sheet)
    _write_df_to_sheet(ws, df_updated)
    wb.save(master_path)

    logger.info(f"Step 6 complete: Added {len(df_new)} new ASINs to master file '{master_sheet}' tab")


def _write_df_to_sheet(ws, df):
    """Write dataframe to openpyxl worksheet with header styling"""
    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    # Auto-width
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


# ─────────────────────────────────────────────
# STEP 7: Save full output Excel with all sheets
# ─────────────────────────────────────────────
def save_output_excel(bot_results, df_suppression, df_instock_removed,
                      df_vlookup, df_less30, df_final, output_path):
    """
    Save output Excel with all intermediate sheets for audit trail
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheets_data = [
        ("bot-run-output",        pd.DataFrame(bot_results)),
        ("Supression",            df_suppression),
        ("Supression(instock removal)", df_instock_removed),
        ("vlookup-sell-unsellable", df_vlookup),
        ("less than 30",          df_less30),
        ("NEW SUPPRESSIONS",      df_final),
    ]

    colors = ["2E86AB", "A23B72", "F18F01", "C73E1D", "3B1F2B", "D62828"]

    for (sheet_name, df), color in zip(sheets_data, colors):
        ws = wb.create_sheet(sheet_name)
        if df is None or len(df) == 0:
            ws.cell(row=1, column=1, value="No data")
            continue

        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
    logger.info(f"Output Excel saved: {output_path}")


# ─────────────────────────────────────────────
# MASTER PIPELINE FUNCTION
# ─────────────────────────────────────────────
def run_pipeline(bot_results, inventory_path, master_path, master_sheet,
                 output_path, min_qty=30):
    """
    Runs the full suppression processing pipeline
    Returns df_final (new suppressions) and summary dict
    """
    summary = {}

    # Step 1: Build suppression sheet
    df_suppression = build_suppression_sheet(bot_results)
    summary['total_scraped'] = len(df_suppression)

    # Step 2: Filter in-stock/healthy ASINs
    df_instock_removed = filter_suppressed_asins(df_suppression)
    summary['after_filter'] = len(df_instock_removed)

    # Step 3: Vlookup sellable qty
    df_vlookup = vlookup_sellable_qty(df_instock_removed, inventory_path)
    summary['after_inventory_match'] = len(df_vlookup)

    # Step 4: Remove < 30 qty
    df_less30 = remove_low_inventory(df_vlookup, min_qty)
    summary['after_qty_filter'] = len(df_less30)

    # Step 5: Find new suppressions
    df_final, df_master = find_new_suppressions(df_less30, master_path, master_sheet)
    summary['new_suppressions'] = len(df_final)

    # Step 6: Update master file
    update_master_file(df_final, master_path, master_sheet)

    # Step 7: Save output Excel
    save_output_excel(
        bot_results, df_suppression, df_instock_removed,
        df_vlookup, df_less30, df_final, output_path
    )

    logger.info(f"\n{'='*50}")
    logger.info(f"PIPELINE SUMMARY:")
    logger.info(f"  Total ASINs scraped   : {summary['total_scraped']}")
    logger.info(f"  After stock filter    : {summary['after_filter']}")
    logger.info(f"  After inventory match : {summary['after_inventory_match']}")
    logger.info(f"  After qty filter (30+): {summary['after_qty_filter']}")
    logger.info(f"  NEW suppressions found: {summary['new_suppressions']}")
    logger.info(f"{'='*50}\n")

    return df_final, summary
